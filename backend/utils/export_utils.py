from typing import Iterable, Dict, List

from io import BytesIO
from xml.sax.saxutils import escape as xml_escape


def _normalize_rows(responses: Iterable[Dict]) -> List[List[str]]:
    rows: List[List[str]] = []
    for r in responses:
        rows.append([
            str(r.get("response_id", "")),
            str(r.get("submitted_at", "")),
            str(r.get("survey_id", "")),
            str(r.get("survey_title", "")),
            str(r.get("question_id", "")),
            str(r.get("question", "")),
            str(r.get("type", "")),
            "" if r.get("rating") is None else str(r.get("rating")),
            str(r.get("choice", "")),
            str(r.get("comment", "")),
        ])
    return rows


def export_responses_to_excel(responses: Iterable[Dict]) -> bytes:
    """Generate an XLSX workbook of responses with basic formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"

    headers = [
        "Response ID",
        "Submitted At",
        "Survey ID",
        "Survey Title",
        "Question ID",
        "Question",
        "Type",
        "Rating",
        "Choice",
        "Comment",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0AD1FF")
    header_font = Font(bold=True, color="0A1F3D")
    center = Alignment(vertical="center")
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in _normalize_rows(responses):
        ws.append(row)

    # Auto width
    for col in ws.columns:
        max_len = 12
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = min(len(val), 60)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_responses_to_pdf(responses: Iterable[Dict]) -> bytes:
    """Generate a simple tabular PDF of responses."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from pathlib import Path

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)

    styles = getSampleStyleSheet()

    font_name = "Helvetica"
    try:
        candidate_paths = []
        # Preferred: project-local font
        candidate_paths.append(Path(__file__).resolve().parent / "fonts" / "NotoSansEthiopic-Regular.ttf")
        # Windows common fonts with Ethiopic/Unicode coverage
        candidate_paths.append(Path("C:/Windows/Fonts/NotoSansEthiopic-Regular.ttf"))
        candidate_paths.append(Path("C:/Windows/Fonts/Nyala.ttf"))
        candidate_paths.append(Path("C:/Windows/Fonts/Ebrima.ttf"))
        candidate_paths.append(Path("C:/Windows/Fonts/arialuni.ttf"))
        # Linux common fonts
        candidate_paths.append(Path("/usr/share/fonts/truetype/noto/NotoSansEthiopic-Regular.ttf"))
        candidate_paths.append(Path("/usr/share/fonts/opentype/noto/NotoSansEthiopic-Regular.ttf"))
        candidate_paths.append(Path("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"))
        candidate_paths.append(Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))

        for p in candidate_paths:
            if p.exists():
                pdfmetrics.registerFont(TTFont("UnifiedUnicode", str(p)))
                font_name = "UnifiedUnicode"
                break
    except Exception:
        font_name = "Helvetica"

    title_style = styles["Title"].clone("TitleUnicode")
    title_style.fontName = font_name
    title = Paragraph("EEU Responses Export", title_style)

    cell_style = styles["BodyText"].clone("CellUnicode")
    cell_style.fontName = font_name
    cell_style.fontSize = 8
    cell_style.leading = 10

    header = [
        "Response ID",
        "Submitted At",
        "Survey ID",
        "Survey Title",
        "Question ID",
        "Question",
        "Type",
        "Rating",
        "Choice",
        "Comment",
    ]
    data = [[Paragraph(xml_escape(h), cell_style) for h in header]]
    for row in _normalize_rows(responses):
        data.append([Paragraph(xml_escape(str(v)), cell_style) for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0, 0.82, 1)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.Color(0.039, 0.122, 0.239)),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))

    try:
        doc.build([title, table])
    except Exception:
        # Fallback: rebuild with ASCII-safe text and core Helvetica to avoid unicode/font errors
        def _ascii_safe(s: str) -> str:
            try:
                return (s or "").encode("latin-1", "ignore").decode("latin-1")
            except Exception:
                return ""

        fallback_style = styles["BodyText"].clone("CellAscii")
        fallback_style.fontName = "Helvetica"
        fallback_style.fontSize = 8
        fallback_style.leading = 10

        header = [
            "Response ID",
            "Submitted At",
            "Survey ID",
            "Survey Title",
            "Question ID",
            "Question",
            "Type",
            "Rating",
            "Choice",
            "Comment",
        ]
        data2 = [[Paragraph(xml_escape(_ascii_safe(h)), fallback_style) for h in header]]
        for row in _normalize_rows(responses):
            data2.append([Paragraph(xml_escape(_ascii_safe(str(v))), fallback_style) for v in row])

        table2 = Table(data2, repeatRows=1)
        table2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0, 0.82, 1)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.Color(0.039, 0.122, 0.239)),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))
            
        try:
            title_fallback = Paragraph("EEU Responses Export", styles["Title"])
        except Exception:
            title_fallback = Paragraph("Responses Export", styles["Title"])
        doc.build([title_fallback, table2])
    return buf.getvalue()
